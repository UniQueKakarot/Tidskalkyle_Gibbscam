""" This script is going to be timecalculating my machiningtime,
    and outputting the results in a excelfile """

import os.path
from pathlib import Path
import sys
import time

import openpyxl as op

# Ugly try and except to just catch anything that comes up and write it to a file (Lazy logging)
try:
    # opening and reading the source file generated by the GibbsCam macro
    # filename = Path("C:\\Users\\heiv085\\Documents\\Github\\Tidskalkyle_Gibbscam\\main\\Time.txt")
    with open('Time.txt', 'r') as f:
        filebody = [line for line in f]

    filebody_clean = [item[:-1] for item in filebody]
    productcode = filebody_clean.pop(0)

    filebody_clean = [float(item) for item in filebody_clean]

    # adding up the time from the source file and rounding it to the nearest int
    time_in_sec = 0.0
    for time in filebody_clean:
        time_in_sec += time


    time_in_sec = int(round(time_in_sec, 0))

    # checking if the excel file exists or not, if not generate
    if os.path.isfile('Tidskalkyle.xlsx') == False:
        wb = op.Workbook()

        wb.create_sheet('Raw Data')
        wb.create_sheet('Results')

        del wb['Sheet']

        ws = wb['Raw Data']
        ws.cell(row=1, column=1, value='Productcode:')
        ws.cell(row=1, column=2, value='Time:')

        ws.cell(row=2, column=1, value=productcode)
        ws.cell(row=2, column=2, value=time_in_sec)

        ws = wb['Results']
        ws.cell(row=1, column=1, value='Productcode:')
        ws.cell(row=1, column=2, value='Total Machining time:')

        wb.save('Tidskalkyle.xlsx')

    else:
        wb = op.load_workbook('Tidskalkyle.xlsx')
        ws = wb['Raw Data']

        row = 2
        count = 0
        while ws.cell(row=row, column=1).value != None:

            if ws.cell(row=row, column=1).value == productcode:

                ws.cell(row=row, column=1, value=productcode)
                ws.cell(row=row, column=2, value=time_in_sec)
                count = 1
            
            row += 1
            
        if count == 0:

            ws.cell(row=row, column=1, value=productcode)
            ws.cell(row=row, column=2, value=time_in_sec)

        wb.save('Tidskalkyle.xlsx')

    #############################################
    # Dealing with the results sheet under      #
    #############################################

    # reading in all product numbers in the raw data sheet into a list
    productnumber = []
    row = 2
    while ws.cell(row=row, column=1).value != None:
        raw_product = ws.cell(row=row, column=1).value
        productnumber.append(raw_product[:6])
        row += 1

    # coverting from a list to a set and back to a list to remove duplicates
    productnumber = list(set(productnumber))

    # main loop for finding all entries with the same number, add up their time
    # and write it out to the results sheet
    time = 0
    row = 2
    for item in productnumber:

        time = 0
        row = 2
        while ws.cell(row=row, column=1).value != None:
            ordernumber = ws.cell(row=row, column=1).value
            ordernumber = ordernumber[:6]

            if ordernumber == item:
                time += ws.cell(row=row, column=2).value

            row += 1

        # switching to the results sheet for summary writing
        ws = wb['Results']

        seconds = 0
        minutes = 0
        hours = 0

        # converting seconds to hours, minutes and seconds
        while time != 0:

            if time > 3600:
                hours += 1
                time -= 3600

            elif time > 60:
                minutes += 1
                time -= 60

            else:
                seconds = time
                time -= time

        total_time = str(hours) + ':' + str(minutes) + ':' + str(seconds)

        # writing the results to the sheet, if number already exists, overwrite time
        # in place
        check = 0
        row_results = 2
        while ws.cell(row=row_results, column=1).value != None:

            if item == ws.cell(row=row_results, column=1).value:
                ws.cell(row=row_results, column=2, value=total_time)
                check = 1
            row_results += 1
        
        if check == 0:

            ws.cell(row=row_results, column=1, value=item)
            ws.cell(row=row_results, column=2, value=total_time)

        row += 1

        # switching back sheet because we are in a loop after all
        ws = wb['Raw Data']

        wb.save('Tidskalkyle.xlsx')
except Exception as e:
    print(e)




input('Enter to end')
